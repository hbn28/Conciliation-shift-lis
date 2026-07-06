import pandas as pd
from openpyxl import load_workbook

from app.adapters.outbound.spreadsheets.reporter import generate_reports
from app.application.models import ComparisonResult


def test_relatorio_detalhado_inclui_auditoria_e_descartes(tmp_path):
    result = ComparisonResult(
        resumo={"total_linhas_rede": 1, "total_linhas_shift": 1},
        detalhado=pd.DataFrame([{"status_comparacao": "CONCILIADO"}]),
        qualidade_shift=pd.DataFrame(),
        auditoria=pd.DataFrame([
            {"origem": "SHIFT", "indicador": "shift_total_linhas_lidas", "valor": 2}
        ]),
        descartes=pd.DataFrame([
            {
                "origem": "SHIFT",
                "numero_linha_original": 7,
                "motivo_descarte": "AUTORIZACAO_AUSENTE",
            }
        ]),
    )
    paths = generate_reports(result, tmp_path)
    workbook = load_workbook(paths["detalhado"], read_only=True)
    assert workbook.sheetnames == ["Detalhado", "Auditoria", "Descartes"]
    assert workbook["Auditoria"]["B2"].value == "shift_total_linhas_lidas"
    assert workbook["Descartes"]["C2"].value == "AUTORIZACAO_AUSENTE"
