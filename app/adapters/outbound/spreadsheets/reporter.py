from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.application.models import ComparisonResult


def _json_default(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, pd.Timestamp)):
        return value.isoformat()
    raise TypeError(f"Tipo não serializável: {type(value)}")


def _serializable_frame(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in result.columns:
        result[col] = result[col].map(
            lambda value: value.isoformat() if isinstance(value, date)
            else float(value) if isinstance(value, Decimal)
            else value
        )
    return result


def _excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    """Evita que textos vindos de upload sejam executados como fórmulas."""
    result = df.copy()
    for col in result.columns:
        result[col] = result[col].map(
            lambda value: f"'{value}"
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@"))
            else value
        )
    return result


def _style_workbook(path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17324D")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for idx, column in enumerate(sheet.columns, 1):
            values = [str(cell.value or "") for cell in list(column)[:200]]
            width = min(max(max(map(len, values), default=8) + 2, 12), 38)
            header = str(sheet.cell(1, idx).value or "")
            if header in {"observacao", "motivo", "sugestao_acao", "acao_recomendada"}:
                width = 60
            sheet.column_dimensions[get_column_letter(idx)].width = width
        sheet.row_dimensions[1].height = 42
    workbook.save(path)


def generate_reports(result: ComparisonResult, output_dir: str | Path) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    detailed = _excel_safe(_serializable_frame(result.detalhado))
    quality = _excel_safe(_serializable_frame(result.qualidade_shift))
    audit = _excel_safe(_serializable_frame(
        result.auditoria if result.auditoria is not None else pd.DataFrame()
    ))
    discards = _excel_safe(_serializable_frame(
        result.descartes if result.descartes is not None else pd.DataFrame()
    ))
    def _summary_value(value):
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, dict):
            # Ex.: quantidade_por_bandeira_shift -> "MASTER: 8, VISA: 3".
            return ", ".join(f"{k}: {v}" for k, v in value.items())
        if isinstance(value, (list, tuple)):
            # Ex.: alertas_empresa_shift -> várias mensagens numa célula.
            return " | ".join(str(v) for v in value) if value else ""
        return value

    summary = pd.DataFrame([
        {"indicador": key, "valor": _summary_value(value)}
        for key, value in result.resumo.items()
    ])
    paths = {
        "detalhado": output_dir / "relatorio_detalhado.xlsx",
        "resumo": output_dir / "resumo.xlsx",
        "qualidade-shift": output_dir / "qualidade_shift.xlsx",
    }
    with pd.ExcelWriter(paths["detalhado"], engine="openpyxl") as writer:
        detailed.to_excel(writer, index=False, sheet_name="Detalhado")
        audit.to_excel(writer, index=False, sheet_name="Auditoria")
        discards.to_excel(writer, index=False, sheet_name="Descartes")
    with pd.ExcelWriter(paths["resumo"], engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Resumo")
    with pd.ExcelWriter(paths["qualidade-shift"], engine="openpyxl") as writer:
        quality.to_excel(writer, index=False, sheet_name="Qualidade Shift")
    for path in paths.values():
        _style_workbook(path)
    detailed.to_csv(output_dir / "relatorio_detalhado.csv", index=False, encoding="utf-8-sig", sep=";")
    quality.to_csv(output_dir / "qualidade_shift.csv", index=False, encoding="utf-8-sig", sep=";")
    (output_dir / "resultado.json").write_text(
        json.dumps({
            "resumo": result.resumo,
            "detalhado": detailed.fillna("").to_dict("records"),
            "qualidade_shift": quality.fillna("").to_dict("records"),
            "auditoria": audit.fillna("").to_dict("records"),
            "descartes": discards.fillna("").to_dict("records"),
        }, ensure_ascii=False, default=_json_default, indent=2),
        encoding="utf-8",
    )
    return paths
